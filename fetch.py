#!/usr/bin/env python3
"""
fetch.py — knowledge-base builder for the ITU AI Readiness contribution.

manifest.csv is the source of truth. Everything else in kb/ is derived and
reproducible from it. Raw copies stay local; the manifest and the derived
text travel in git.

Subcommands
-----------
  validate   check manifest.csv against manifest.schema.json
  fetch      download documents, hash them, record provenance
  verify     re-hash held copies and re-check upstream for drift
  extract    PDF/HTML -> plain text
  chunk      text -> chunks.jsonl with character offsets and page numbers
  report     coverage by ITU factor / dimension / status / sector
  export     write manifest.xlsx for human review

Typical first run:
  python fetch.py validate
  python fetch.py fetch
  python fetch.py extract && python fetch.py chunk
  python fetch.py report

Dependencies: requests, beautifulsoup4, pdfplumber (or pypdf).
Optional: jsonschema (a built-in fallback validator is used if absent),
          openpyxl (only for `export`).
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
import time
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

KB = Path(__file__).resolve().parent
MANIFEST = KB / "manifest.csv"
SCHEMA = KB / "manifest.schema.json"
RAW = KB / "raw"
TEXT = KB / "processed" / "text"
CHUNKS = KB / "processed" / "chunks.jsonl"

# Columns holding ';'-separated lists, and their element type.
LIST_COLUMNS = {"itu_factors": str, "itu_dimensions": int}
BOOL_COLUMNS = {"is_official_translation"}
INT_COLUMNS = {"year", "http_status", "page_count", "bytes"}

USER_AGENT = (
    "ITU-AIReadiness-KB-Collector/1.0 "
    "(academic hackathon research; contact: <your-email>)"
)
TIMEOUT = 45
RETRIES = 3
POLITE_DELAY = 1.5  # seconds between requests to the same host


# --------------------------------------------------------------------------
# manifest I/O
# --------------------------------------------------------------------------
def load_manifest(path: Path = MANIFEST) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Read manifest.csv into typed dicts. Returns (rows, fieldnames)."""
    if not path.exists():
        sys.exit(f"ERROR: {path} not found.")
    with path.open(encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        fieldnames = list(reader.fieldnames or [])
        rows = [_decode_row(r) for r in reader]
    return rows, fieldnames


def _decode_row(raw: Dict[str, str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for key, val in raw.items():
        if key is None:
            continue
        val = (val or "").strip()
        if val == "":
            continue
        if key in LIST_COLUMNS:
            caster = LIST_COLUMNS[key]
            items = [p.strip() for p in val.split(";") if p.strip()]
            try:
                out[key] = [caster(p) for p in items]
            except ValueError:
                out[key] = items  # let the validator complain, not the parser
        elif key in BOOL_COLUMNS:
            out[key] = val.lower() in {"true", "yes", "1", "y"}
        elif key in INT_COLUMNS:
            try:
                # Excel writes whole numbers as "2025.0"; accept both forms.
                out[key] = int(float(val))
            except ValueError:
                out[key] = val
        else:
            out[key] = val
    return out


def _encode_row(row: Dict[str, Any], fieldnames: List[str]) -> Dict[str, str]:
    out = {}
    for key in fieldnames:
        val = row.get(key, "")
        if isinstance(val, list):
            val = ";".join(str(v) for v in val)
        elif isinstance(val, bool):
            val = "true" if val else "false"
        out[key] = "" if val is None else str(val)
    return out


def save_manifest(rows, fieldnames, path: Path = MANIFEST) -> None:
    """Rewrite manifest.csv, preserving column order. Writes atomically."""
    tmp = path.with_suffix(".csv.tmp")
    with tmp.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(_encode_row(row, fieldnames))
    tmp.replace(path)


def by_tier(rows: List[Dict[str, Any]], tier: Optional[str]) -> List[Dict[str, Any]]:
    """Filter rows by tier. `--tier core` is the eight-document demo core."""
    if not tier or tier == "all":
        return rows
    return [r for r in rows if r.get("tier") == tier]


# --------------------------------------------------------------------------
# validation
# --------------------------------------------------------------------------
def _fallback_validate(row: Dict[str, Any], schema: Dict[str, Any]) -> List[str]:
    """Minimal validator covering required / enum / pattern / type / range.

    Used when jsonschema is unavailable. Deliberately strict about the things
    that actually break downstream work (doc_id format, enums, hash format).
    """
    errs: List[str] = []
    props = schema.get("properties", {})

    for req in schema.get("required", []):
        if req not in row or row[req] in ("", None, []):
            errs.append(f"missing required field '{req}'")

    for key, val in row.items():
        spec = props.get(key)
        if spec is None:
            errs.append(f"unknown column '{key}' (not in schema)")
            continue
        errs.extend(_check_value(key, val, spec))
    return errs


def _check_value(key: str, val: Any, spec: Dict[str, Any]) -> List[str]:
    errs: List[str] = []
    typ = spec.get("type")

    if typ == "array":
        if not isinstance(val, list):
            return [f"'{key}': expected list, got {type(val).__name__}"]
        item_spec = spec.get("items", {})
        for item in val:
            errs.extend(_check_value(f"{key}[]", item, item_spec))
        if spec.get("uniqueItems") and len(set(map(str, val))) != len(val):
            errs.append(f"'{key}': duplicate entries")
        return errs

    if typ == "integer":
        if not isinstance(val, int) or isinstance(val, bool):
            return [f"'{key}': expected integer, got {val!r}"]
        if "minimum" in spec and val < spec["minimum"]:
            errs.append(f"'{key}': {val} below minimum {spec['minimum']}")
        if "maximum" in spec and val > spec["maximum"]:
            errs.append(f"'{key}': {val} above maximum {spec['maximum']}")
        return errs

    if typ == "boolean":
        if not isinstance(val, bool):
            errs.append(f"'{key}': expected boolean, got {val!r}")
        return errs

    if typ == "string":
        if not isinstance(val, str):
            return [f"'{key}': expected string, got {val!r}"]
        if "enum" in spec and val not in spec["enum"]:
            errs.append(f"'{key}': '{val}' not in {spec['enum']}")
        if "pattern" in spec and not re.match(spec["pattern"], val):
            errs.append(f"'{key}': '{val}' fails pattern {spec['pattern']}")
        if "minLength" in spec and len(val) < spec["minLength"]:
            errs.append(f"'{key}': shorter than {spec['minLength']} chars")
        if spec.get("format") == "date" and not re.match(r"^\d{4}-\d{2}-\d{2}$", val):
            errs.append(f"'{key}': '{val}' is not an ISO date")
    return errs


def cmd_validate(args) -> int:
    schema = json.loads(SCHEMA.read_text(encoding="utf-8"))
    all_rows, _ = load_manifest()
    rows = by_tier(all_rows, getattr(args, "tier", None))

    try:
        import jsonschema  # type: ignore

        validator = jsonschema.Draft202012Validator(schema)

        def run(row):
            return [e.message for e in validator.iter_errors(row)]

        engine = "jsonschema"
    except ImportError:
        def run(row):
            return _fallback_validate(row, schema)

        engine = "built-in fallback (pip install jsonschema for full coverage)"

    print(f"Validating {len(rows)} records with {engine}\n")
    bad = 0
    seen_ids: Dict[str, int] = {}

    for i, row in enumerate(rows, start=2):  # +1 header, +1 to 1-index
        errs = run(row)
        did = row.get("doc_id", f"<row {i}>")
        if did in seen_ids:
            errs.append(f"duplicate doc_id, first seen on line {seen_ids[did]}")
        else:
            seen_ids[did] = i
        if errs:
            bad += 1
            print(f"  line {i}  {did}")
            for e in errs:
                print(f"      - {e}")

    # Soft warnings: things that are legal but will cost marks.
    warn = [r for r in rows if not r.get("url")]
    if warn:
        core_left = [r for r in warn if r.get("tier") == "core"]
        print(f"\n  NOTE: {len(warn)} record(s) have no url yet"
              f" ({len(core_left)} of them CORE):")
        for r in sorted(warn, key=lambda x: (x.get("tier") != "core", x.get("doc_id", ""))):
            flag = "CORE " if r.get("tier") == "core" else "     "
            print(f"      - {flag}{r.get('doc_id')}  {r.get('title_en','')[:55]}")
        print("      Source these from the issuing body's own domain. Do not")
        print("      copy URLs from aggregators, and do not invent them.")

    unk = [r for r in rows if r.get("status") == "unknown"]
    if unk:
        print(f"\n  NOTE: {len(unk)} record(s) have status='unknown'. Resolve")
        print("      binding vs advisory before submission — gap analysis that")
        print("      conflates the two is not defensible.")

    print(f"\n{'FAIL' if bad else 'PASS'}: {bad} invalid record(s).")
    return 1 if bad else 0


# --------------------------------------------------------------------------
# fetching
# --------------------------------------------------------------------------
def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _guess_format(content_type: str, url: str, data: bytes) -> str:
    ct = (content_type or "").lower()
    if data[:5] == b"%PDF-":
        return "pdf"
    if "pdf" in ct or url.lower().endswith(".pdf"):
        return "pdf"
    if "html" in ct or url.lower().endswith((".html", ".htm", "/")):
        return "html"
    if "openxmlformats-officedocument.wordprocessing" in ct or url.lower().endswith(".docx"):
        return "docx"
    if "spreadsheet" in ct or url.lower().endswith((".xlsx", ".xls")):
        return "xlsx"
    if "text/plain" in ct:
        return "txt"
    return "unknown"


def _download(url: str) -> Tuple[Optional[bytes], int, str]:
    import requests

    last = ""
    for attempt in range(1, RETRIES + 1):
        try:
            resp = requests.get(
                url,
                headers={"User-Agent": USER_AGENT, "Accept": "*/*"},
                timeout=TIMEOUT,
                allow_redirects=True,
            )
            if resp.status_code == 200:
                return resp.content, 200, resp.headers.get("Content-Type", "")
            last = f"HTTP {resp.status_code}"
            if resp.status_code in (401, 403, 404, 410):
                return None, resp.status_code, ""  # not worth retrying
        except Exception as exc:  # noqa: BLE001 — report, don't crash the run
            last = f"{type(exc).__name__}: {exc}"
        if attempt < RETRIES:
            time.sleep(2 * attempt)
    print(f"      failed after {RETRIES} attempts: {last}")
    return None, 0, ""


def _page_count(path: Path) -> int:
    try:
        from pypdf import PdfReader

        return len(PdfReader(str(path)).pages)
    except Exception:
        return 0


def cmd_fetch(args) -> int:
    rows, fieldnames = load_manifest()
    RAW.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()

    targets = [r for r in by_tier(rows, getattr(args, "tier", None)) if r.get("url")]
    if args.only:
        wanted = set(args.only)
        targets = [r for r in targets if r.get("doc_id") in wanted]

    print(f"Fetching {len(targets)} document(s)\n")
    ok = failed = skipped = drifted = 0

    for row in targets:
        did, url = row["doc_id"], row["url"]
        existing = sorted(RAW.glob(f"{did}.*"))
        if existing and not args.force:
            print(f"  SKIP    {did}  (already held: {existing[0].name})")
            skipped += 1
            continue

        print(f"  GET     {did}")
        data, status, ctype = _download(url)
        row["http_status"] = status
        if data is None:
            failed += 1
            continue

        fmt = _guess_format(ctype, url, data)
        digest = _sha256(data)
        prior = row.get("sha256")
        if prior and prior != digest:
            # This is the corpus-level version of the drift the project studies.
            print(f"      DRIFT: upstream copy differs from recorded sha256")
            print(f"             was {prior[:16]}...  now {digest[:16]}...")
            drifted += 1

        dest = RAW / f"{did}.{fmt if fmt != 'unknown' else 'bin'}"
        dest.write_bytes(data)

        row.update(
            {
                "sha256": digest,
                "bytes": len(data),
                "format": fmt,
                "url_accessed": today,
                "local_path": str(dest.relative_to(KB)),
            }
        )
        if fmt == "pdf":
            row["page_count"] = _page_count(dest)
        print(f"      saved {dest.name}  {len(data):,} bytes  sha256 {digest[:16]}...")
        ok += 1
        time.sleep(POLITE_DELAY)

    save_manifest(rows, fieldnames)
    print(
        f"\nfetched {ok} | skipped {skipped} | failed {failed} | drifted {drifted}"
    )
    if failed:
        print("Re-source failed URLs from the issuing body's site by hand.")
    return 0


def cmd_verify(args) -> int:
    """Re-hash local copies; optionally re-fetch to detect upstream revision."""
    rows, fieldnames = load_manifest()

    if getattr(args, "rehash", False):
        print("Re-hashing held copies (on-disk file treated as authoritative)\n")
        done = 0
        for row in rows:
            lp = row.get("local_path")
            if not lp:
                continue
            path = KB / lp
            if not path.exists():
                print(f"  MISSING  {row.get('doc_id')}  ({lp})")
                continue
            data = path.read_bytes()
            digest = _sha256(data)
            prior = row.get("sha256")
            row["sha256"] = digest
            row["bytes"] = len(data)
            if row.get("format") == "pdf":
                row["page_count"] = _page_count(path)
            state = "unchanged" if prior == digest else ("updated" if prior else "new")
            print(f"  {state:<9} {row.get('doc_id'):<26} "
                  f"{len(data):>12,} bytes  {digest[:16]}...  "
                  f"p{row.get('page_count') or '-'}")
            done += 1
        save_manifest(rows, fieldnames)
        print(f"\nre-hashed {done} record(s); manifest updated")
        return 0

    print("Verifying held copies against recorded hashes\n")
    local_bad = missing = upstream_changed = 0

    for row in rows:
        did = row.get("doc_id")
        lp = row.get("local_path")
        recorded = row.get("sha256")
        if not lp or not recorded:
            continue
        path = KB / lp
        if not path.exists():
            print(f"  MISSING  {did}  ({lp})")
            missing += 1
            continue
        actual = _sha256(path.read_bytes())
        if actual != recorded:
            print(f"  ALTERED  {did}  local file no longer matches manifest")
            local_bad += 1

        if args.upstream and row.get("url"):
            data, _, _ = _download(row["url"])
            time.sleep(POLITE_DELAY)
            if data is not None and _sha256(data) != recorded:
                print(f"  UPSTREAM {did}  publisher has revised this document")
                upstream_changed += 1

    print(
        f"\naltered {local_bad} | missing {missing} | upstream revised {upstream_changed}"
    )
    if upstream_changed:
        print("Re-run `fetch --force --only <doc_id>` and re-run the gap analysis.")
    return 0


# --------------------------------------------------------------------------
# text extraction
# --------------------------------------------------------------------------
def _pdf_to_pages(path: Path) -> List[str]:
    try:
        import pdfplumber

        with pdfplumber.open(str(path)) as pdf:
            return [(p.extract_text() or "") for p in pdf.pages]
    except Exception:
        pass
    try:
        from pypdf import PdfReader

        return [(p.extract_text() or "") for p in PdfReader(str(path)).pages]
    except Exception as exc:  # noqa: BLE001
        print(f"      extraction failed: {exc}")
        return []


def _html_to_text(path: Path) -> List[str]:
    from bs4 import BeautifulSoup

    soup = BeautifulSoup(path.read_bytes(), "lxml")
    for tag in soup(["script", "style", "nav", "header", "footer", "noscript", "form"]):
        tag.decompose()
    main = soup.find("main") or soup.find("article") or soup.body or soup
    text = main.get_text("\n")
    lines = [ln.strip() for ln in text.splitlines()]
    return ["\n".join(ln for ln in lines if ln)]


def cmd_extract(args) -> int:
    rows, _ = load_manifest()
    rows = by_tier(rows, getattr(args, "tier", None))
    TEXT.mkdir(parents=True, exist_ok=True)
    done = empty = 0

    for row in rows:
        lp = row.get("local_path")
        if not lp:
            continue
        path = KB / lp
        if not path.exists():
            continue
        did, fmt = row["doc_id"], row.get("format", "unknown")

        if fmt == "pdf":
            pages = _pdf_to_pages(path)
        elif fmt == "html":
            pages = _html_to_text(path)
        elif fmt in ("json", "xlsx", "docx"):
            print(f"  SKIP  {did}  ({fmt}: structured source, parsed by the")
            print(f"              gap engine, not text-chunked — this is correct)")
            continue
        else:
            print(f"  SKIP  {did}  (format '{fmt}' not handled)")
            continue

        body = "\n\n".join(f"<<<PAGE {i}>>>\n{t}" for i, t in enumerate(pages, 1))
        if not body.strip() or sum(len(p) for p in pages) < 200:
            # Almost always a scanned document with no text layer.
            print(f"  EMPTY {did}  — likely scanned; needs OCR before use")
            empty += 1
            continue
        (TEXT / f"{did}.txt").write_text(body, encoding="utf-8")
        print(f"  OK    {did}  {len(pages)} page(s), {len(body):,} chars")
        done += 1

    print(f"\nextracted {done} | empty/scanned {empty}")
    return 0


# --------------------------------------------------------------------------
# chunking
# --------------------------------------------------------------------------
PAGE_RE = re.compile(r"<<<PAGE (\d+)>>>\n")


def _split_pages(text: str) -> List[Tuple[int, str]]:
    parts = PAGE_RE.split(text)
    out: List[Tuple[int, str]] = []
    for i in range(1, len(parts) - 1, 2):
        out.append((int(parts[i]), parts[i + 1]))
    return out or [(1, text)]


def _chunk_page(body: str, size: int, overlap: int) -> Iterable[Tuple[int, str]]:
    """Yield (start_offset, text). Breaks on paragraph, then sentence, then hard.

    Offsets are relative to the page body, so every chunk stays traceable back
    to an exact span of the source document.
    """
    n = len(body)
    if n == 0:
        return
    if n <= size:
        piece = body.strip()
        if piece:
            yield 0, piece
        return

    i = 0
    while i < n:
        end = min(i + size, n)
        if end < n:
            window = body[i:end]
            for sep in ("\n\n", "\n", ". ", " "):
                cut = window.rfind(sep)
                if cut > size * 0.5:
                    end = i + cut + len(sep)
                    break
        piece = body[i:end].strip()
        if piece:
            yield i, piece
        if end >= n:
            break
        nxt = end - overlap
        if nxt <= i:  # guard against non-advancing windows
            nxt = i + max(1, size // 2)
        i = nxt


def cmd_chunk(args) -> int:
    rows, _ = load_manifest()
    keep = {r["doc_id"] for r in by_tier(rows, getattr(args, "tier", None))}
    by_id = {r["doc_id"]: r for r in rows if r.get("doc_id")}
    CHUNKS.parent.mkdir(parents=True, exist_ok=True)
    total = 0

    with CHUNKS.open("w", encoding="utf-8") as out:
        for txt_path in sorted(TEXT.glob("*.txt")):
            did = txt_path.stem
            if did not in keep:
                continue
            meta = by_id.get(did, {})
            content = txt_path.read_text(encoding="utf-8")
            k = 0
            for page_no, page_body in _split_pages(content):
                for start, piece in _chunk_page(page_body, args.size, args.overlap):
                    k += 1
                    out.write(
                        json.dumps(
                            {
                                # chunk_id is the citation anchor the app must
                                # print with every answer.
                                "chunk_id": f"{did}::p{page_no}::c{k:04d}",
                                "doc_id": did,
                                "page": page_no,
                                "char_start": start,
                                "char_end": start + len(piece),
                                "text": piece,
                                "title_en": meta.get("title_en", ""),
                                "issuing_body": meta.get("issuing_body", ""),
                                "status": meta.get("status", ""),
                                "sector": meta.get("sector", ""),
                                "year": meta.get("year", ""),
                                "url": meta.get("url", ""),
                                "sha256": meta.get("sha256", ""),
                            },
                            ensure_ascii=False,
                        )
                        + "\n"
                    )
            print(f"  {did}: {k} chunk(s)")
            total += k

    print(f"\nwrote {total} chunks -> {CHUNKS.relative_to(KB.parent)}")
    return 0


# --------------------------------------------------------------------------
# reporting
# --------------------------------------------------------------------------
FACTORS = ["data", "research", "deployment_support", "standards", "open_source", "sandbox"]


def cmd_report(args) -> int:
    all_rows, _ = load_manifest()
    rows = by_tier(all_rows, getattr(args, "tier", None))

    # Methodological references (ITU, OLIR) are not policy content and must not
    # inflate the knowledge-base size claimed in the submission.
    content = [r for r in rows if r.get("tier") != "method"]
    held = [r for r in content if r.get("sha256")]

    print("=" * 62)
    print("KNOWLEDGE BASE COVERAGE")
    print("=" * 62)

    core = [r for r in all_rows if r.get("tier") == "core"]
    core_done = [r for r in core if r.get("sha256")]
    bar_n = len(core_done)
    bar = "#" * bar_n + "." * (len(core) - bar_n)
    print(f"CORE PROGRESS       : [{bar}] {bar_n}/{len(core)} collected")
    if bar_n < len(core):
        print("                      finish core before touching extended")

    print(f"\npolicy documents    : {len(content)} in manifest, {len(held)} held")
    print(f"pages held          : {sum(int(r.get('page_count') or 0) for r in held):,}")
    meth = [r for r in rows if r.get("tier") == "method"]
    if meth:
        print(f"method references   : {len(meth)} (excluded from KB size)")

    def tally(field: str, title: str):
        c = Counter(r.get(field, "—") for r in content)
        print(f"\n{title}")
        for key, num in c.most_common():
            print(f"  {str(key):<24} {num:>3}")

    tally("status", "By legal status")
    tally("sector", "By sector")
    tally("doc_type", "By document type")
    tally("issuing_body", "By issuing body")

    print("\nITU factors covered")
    fc: Counter = Counter()
    for r in content:
        for f in r.get("itu_factors", []) or []:
            fc[f] += 1
    for f in FACTORS:
        mark = " " if fc.get(f) else "  <-- GAP"
        print(f"  {f:<24} {fc.get(f, 0):>3}{mark}")

    print("\nITU dimensions covered (1-13)")
    dc: Counter = Counter()
    for r in content:
        for d in r.get("itu_dimensions", []) or []:
            dc[d] += 1
    missing = [d for d in range(1, 14) if not dc.get(d)]
    line = "  " + " ".join(f"{d:>2}:{dc.get(d,0)}" for d in range(1, 14))
    print(line)
    if missing:
        print(f"  uncovered dimensions: {missing}")
        print("  (Not every dimension must be covered — but be able to say why.)")

    pairs = [r for r in content if r.get("reference_pair")]
    print(f"\nnational<->international pairs configured: {len(pairs)}")
    for r in pairs:
        print(f"  {r['doc_id']}  ->  {r['reference_pair']}")

    todo = [r for r in content if not r.get("url")]
    if todo:
        core_todo = [r for r in todo if r.get("tier") == "core"]
        print(f"\nstill needing a URL: {len(todo)}  (CORE: {len(core_todo)})")
        for r in core_todo:
            print(f"  CORE  {r['doc_id']}")
    return 0


def cmd_export(args) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl not installed:  pip install openpyxl")

    rows, fieldnames = load_manifest()
    wb = Workbook()
    ws = wb.active
    ws.title = "manifest"
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        enc = _encode_row(row, fieldnames)
        ws.append([enc[f] for f in fieldnames])
    ws.freeze_panes = "A2"
    for i, name in enumerate(fieldnames, 1):
        width = max(len(name) + 2, min(48, max((len(str(r.get(name, ""))) for r in rows), default=10) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width
    dest = KB / "manifest.xlsx"
    wb.save(dest)
    print(f"wrote {dest}")
    print("Reminder: manifest.csv stays the source of truth. This copy is")
    print("for reading and review only — edits here are not picked up.")
    return 0


# --------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    TIERS = ["core", "extended", "method", "all"]

    p = sub.add_parser("validate", help="check manifest.csv against the schema")
    p.add_argument("--tier", choices=TIERS, help="limit to one tier")

    p = sub.add_parser("fetch", help="download documents and record provenance")
    p.add_argument("--force", action="store_true", help="re-download held copies")
    p.add_argument("--only", nargs="+", metavar="DOC_ID", help="limit to these doc_ids")
    p.add_argument("--tier", choices=TIERS, help="limit to one tier (start with core)")

    p = sub.add_parser("verify", help="re-hash held copies; detect drift")
    p.add_argument("--upstream", action="store_true", help="also re-fetch and compare")
    p.add_argument("--rehash", action="store_true", help="adopt on-disk files as authoritative")

    p = sub.add_parser("extract", help="PDF/HTML -> plain text")
    p.add_argument("--tier", choices=TIERS, help="limit to one tier")

    p = sub.add_parser("chunk", help="text -> chunks.jsonl")
    p.add_argument("--size", type=int, default=1200, help="target chunk chars (default 1200)")
    p.add_argument("--overlap", type=int, default=150, help="overlap chars (default 150)")
    p.add_argument("--tier", choices=TIERS, help="limit to one tier")

    p = sub.add_parser("report", help="coverage report")
    p.add_argument("--tier", choices=TIERS, help="limit to one tier")
    sub.add_parser("export", help="write manifest.xlsx for humans")

    args = ap.parse_args()
    return {
        "validate": cmd_validate,
        "fetch": cmd_fetch,
        "verify": cmd_verify,
        "extract": cmd_extract,
        "chunk": cmd_chunk,
        "report": cmd_report,
        "export": cmd_export,
    }[args.cmd](args)


if __name__ == "__main__":
    sys.exit(main())
